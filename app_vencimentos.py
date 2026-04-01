import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import io
import re
import zipfile
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from historico_db import init_db, save_run, list_runs, get_run, excluir_run

# Configuração da página
st.set_page_config(
    page_title="Acompanhamento: Comparação de Bases por Ano de Vencimento",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizado
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f4e79;
        text-align: center;
        padding: 1rem 0;
        margin-bottom: 2rem;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #1f4e79;
    }
    .stButton>button {
        width: 100%;
        background-color: #1f4e79;
        color: white;
        font-weight: bold;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 0.5rem;
        padding: 1rem;
        margin: 1rem 0;
    }
    .warning-box {
        background-color: #fff3cd;
        border: 1px solid #ffeaa7;
        border-radius: 0.5rem;
        padding: 1rem;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# Título principal
st.markdown('<div class="main-header">📅 Acompanhamento: Comparação de Bases por Ano de Vencimento</div>', unsafe_allow_html=True)
st.markdown('<p style="text-align: center; color: #666; font-size: 1.1rem;">Compare duas bases e veja quantos autos saíram ou entraram por ano de vencimento</p>', unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.image("https://via.placeholder.com/200x80/1f4e79/ffffff?text=ANTT", use_container_width=True)
    st.markdown("### 📁 Upload das Bases")
    
    st.markdown("#### Base antiga (período anterior)")
    arquivo_antiga = st.file_uploader(
        "Planilha base antiga",
        type=['xlsx', 'xls', 'csv'],
        key='arquivo_base_antiga'
    )
    
    st.markdown("#### Base nova (período atual)")
    arquivo_nova = st.file_uploader(
        "Planilha base nova",
        type=['xlsx', 'xls', 'csv'],
        key='arquivo_base_nova'
    )
    
    st.markdown("---")
    st.markdown("### ⚙️ Configurações")
    
    st.markdown("#### 🔑 Colunas Obrigatórias")
    # Campo para identificar Auto de Infração
    coluna_auto = st.text_input(
        "Nome da coluna Auto de Infração",
        value="Identificador do Débito",
        help="⚠️ OBRIGATÓRIO: Digite o nome exato da coluna que contém os Autos de Infração"
    )
    
    # Campo para identificar data de vencimento
    coluna_vencimento = st.text_input(
        "Nome da coluna Vencimento",
        value="Data do Vencimento",
        help="⚠️ OBRIGATÓRIO: Digite o nome exato da coluna que contém as datas de vencimento"
    )
    
    st.markdown("---")
    st.markdown("#### 📋 Colunas Opcionais")
    # Campo para identificar número de protocolos (opcional)
    coluna_protocolo = st.text_input(
        "Nome da coluna Nº do Processo (Opcional)",
        value="Nº do Processo",
        help="Digite o nome exato da coluna que contém os números de protocolos (opcional)"
    )
    
    # Campo para identificar Subtipo de Débito (Modal) (opcional)
    coluna_modal = st.text_input(
        "Nome da coluna Subtipo de Débito (Opcional)",
        value="Subtipo de Débito",
        help="Digite o nome exato da coluna que contém os modais (opcional)"
    )
    
    # Campo para identificar CPF/CNPJ (opcional)
    coluna_cpf_cnpj = st.text_input(
        "Nome da coluna CPF/CNPJ (Opcional)",
        value="CPF/CNPJ",
        help="Digite o nome exato da coluna que contém CPF/CNPJ (opcional)"
    )

# Função para carregar dados
@st.cache_data
def carregar_dados(arquivo):
    try:
        if arquivo.name.endswith('.csv'):
            df = pd.read_csv(arquivo, encoding='utf-8', sep=';', decimal=',', header=0)
        else:
            df = pd.read_excel(arquivo, header=0)
        
        # Remover linhas completamente vazias
        df = df.dropna(how='all')
        
        return df
    except Exception as e:
        st.error(f"Erro ao carregar arquivo: {str(e)}")
        return None

# Função para remover duplicados mantendo data mais antiga
def remover_duplicados_manter_mais_antiga(df, coluna_auto, coluna_vencimento):
    """Remove duplicados baseado em Auto de Infração, mantendo a data de vencimento mais antiga"""
    df_resultado = df.copy()
    
    # Converter data de vencimento para datetime
    try:
        if df_resultado[coluna_vencimento].dtype != 'datetime64[ns]':
            df_resultado['_VENCIMENTO_DT'] = pd.to_datetime(
                df_resultado[coluna_vencimento],
                errors='coerce',
                dayfirst=True,
                infer_datetime_format=True
            )
        else:
            df_resultado['_VENCIMENTO_DT'] = df_resultado[coluna_vencimento]
        
        # Ordenar por Auto de Infração e depois por data (mais antiga primeiro)
        df_resultado = df_resultado.sort_values(
            by=[coluna_auto, '_VENCIMENTO_DT'],
            ascending=[True, True],  # Mais antiga primeiro
            na_position='last'
        )
        
        # Remover duplicados mantendo a primeira ocorrência (que será a mais antiga)
        df_resultado = df_resultado.drop_duplicates(
            subset=[coluna_auto],
            keep='first'
        ).copy()
        
        # Remover coluna auxiliar
        if '_VENCIMENTO_DT' in df_resultado.columns:
            df_resultado = df_resultado.drop(columns=['_VENCIMENTO_DT'])
        
        return df_resultado
    except Exception as e:
        st.error(f"Erro ao remover duplicados: {str(e)}")
        return df

# Função para extrair ano da data de vencimento
def extrair_ano_vencimento(df, coluna_vencimento):
    """Extrai o ano da coluna de vencimento"""
    df_resultado = df.copy()
    
    # Converter para datetime
    try:
        if df_resultado[coluna_vencimento].dtype != 'datetime64[ns]':
            df_resultado['_VENCIMENTO_DT'] = pd.to_datetime(
                df_resultado[coluna_vencimento],
                errors='coerce',
                dayfirst=True,
                infer_datetime_format=True
            )
        else:
            df_resultado['_VENCIMENTO_DT'] = df_resultado[coluna_vencimento]
        
        # Extrair ano
        df_resultado['ANO_VENCIMENTO'] = df_resultado['_VENCIMENTO_DT'].dt.year
        
        # Remover coluna auxiliar
        if '_VENCIMENTO_DT' in df_resultado.columns:
            df_resultado = df_resultado.drop(columns=['_VENCIMENTO_DT'])
        
        return df_resultado
    except Exception as e:
        st.error(f"Erro ao extrair ano: {str(e)}")
        return df

# Função para comparar duas bases e calcular saíram/entraram por ano
def comparar_bases(df_antiga, df_nova, coluna_auto, coluna_vencimento):
    """
    Processa base antiga e nova (remove duplicados, extrai ano).
    Para cada ano de vencimento: quantidade na antiga, na nova, quantos saíram, quantos entraram.
    Retorna: comparacao_df, autos_sairam_por_ano, autos_entraram_por_ano, df_antiga_com_ano, df_nova_com_ano,
             stats_antiga_por_ano, stats_nova_por_ano, anos_todos
    """
    # Processar base antiga
    df_antiga_limpa = remover_duplicados_manter_mais_antiga(df_antiga, coluna_auto, coluna_vencimento)
    df_antiga_com_ano = extrair_ano_vencimento(df_antiga_limpa, coluna_vencimento)
    df_antiga_com_ano = df_antiga_com_ano[df_antiga_com_ano['ANO_VENCIMENTO'].notna()].copy()
    df_antiga_com_ano['ANO_VENCIMENTO'] = df_antiga_com_ano['ANO_VENCIMENTO'].astype(int)

    # Processar base nova
    df_nova_limpa = remover_duplicados_manter_mais_antiga(df_nova, coluna_auto, coluna_vencimento)
    df_nova_com_ano = extrair_ano_vencimento(df_nova_limpa, coluna_vencimento)
    df_nova_com_ano = df_nova_com_ano[df_nova_com_ano['ANO_VENCIMENTO'].notna()].copy()
    df_nova_com_ano['ANO_VENCIMENTO'] = df_nova_com_ano['ANO_VENCIMENTO'].astype(int)

    # Conjuntos de autos por ano (identificador normalizado como string)
    def autos_por_ano(df_com_ano):
        d = {}
        for ano, g in df_com_ano.groupby('ANO_VENCIMENTO'):
            d[int(ano)] = set(g[coluna_auto].astype(str).str.strip().values)
        return d

    antiga_por_ano = autos_por_ano(df_antiga_com_ano)
    nova_por_ano = autos_por_ano(df_nova_com_ano)
    anos_todos = sorted(set(antiga_por_ano.keys()) | set(nova_por_ano.keys()))

    # Estatísticas por ano para cada base (para exportação/visualização)
    stats_antiga_por_ano = {}
    for ano in anos_todos:
        df_ano = df_antiga_com_ano[df_antiga_com_ano['ANO_VENCIMENTO'] == ano]
        stats_antiga_por_ano[ano] = {'quantidade': len(df_ano), 'dataframe': df_ano}
    stats_nova_por_ano = {}
    for ano in anos_todos:
        df_ano = df_nova_com_ano[df_nova_com_ano['ANO_VENCIMENTO'] == ano]
        stats_nova_por_ano[ano] = {'quantidade': len(df_ano), 'dataframe': df_ano}

    # Comparação: saíram = estavam na antiga naquele ano e não estão na nova; entraram = estão na nova e não estavam na antiga
    linhas = []
    autos_sairam_por_ano = {}
    autos_entraram_por_ano = {}
    for ano in anos_todos:
        set_antiga = antiga_por_ano.get(ano, set())
        set_nova = nova_por_ano.get(ano, set())
        sairam = set_antiga - set_nova
        entraram = set_nova - set_antiga
        autos_sairam_por_ano[ano] = sairam
        autos_entraram_por_ano[ano] = entraram
        linhas.append({
            'Ano': ano,
            'Qtd base antiga': len(set_antiga),
            'Qtd base nova': len(set_nova),
            'Sairam': len(sairam),
            'Entraram': len(entraram),
        })
    comparacao_df = pd.DataFrame(linhas)

    return {
        'comparacao_df': comparacao_df,
        'autos_sairam_por_ano': autos_sairam_por_ano,
        'autos_entraram_por_ano': autos_entraram_por_ano,
        'df_antiga_com_ano': df_antiga_com_ano,
        'df_nova_com_ano': df_nova_com_ano,
        'stats_antiga_por_ano': stats_antiga_por_ano,
        'stats_nova_por_ano': stats_nova_por_ano,
        'anos_todos': anos_todos,
        'duplicados_antiga': len(df_antiga) - len(df_antiga_limpa),
        'duplicados_nova': len(df_nova) - len(df_nova_limpa),
    }

# Função para formatar CPF/CNPJ no formato brasileiro
def formatar_cpf_cnpj_brasileiro(valor):
    """Formata CPF/CNPJ no formato brasileiro"""
    if pd.isna(valor) or valor == '' or valor is None:
        return ''
    
    # Remove caracteres não numéricos
    valor_str = str(valor).replace('.', '').replace('-', '').replace('/', '').strip()
    
    # Se estiver vazio, retorna vazio
    if not valor_str or not valor_str.isdigit():
        return str(valor)
    
    # Formatar CPF (11 dígitos)
    if len(valor_str) == 11:
        return f"{valor_str[0:3]}.{valor_str[3:6]}.{valor_str[6:9]}-{valor_str[9:11]}"
    
    # Formatar CNPJ (14 dígitos)
    elif len(valor_str) == 14:
        return f"{valor_str[0:2]}.{valor_str[2:5]}.{valor_str[5:8]}/{valor_str[8:12]}-{valor_str[12:14]}"
    
    # Se não tiver 11 ou 14 dígitos, retorna o valor original
    return str(valor)

# Função para gerar Excel formatado (mesmo estilo do app.py)
def gerar_excel_formatado(dados_df, nome_aba, nome_arquivo):
    """Gera arquivo Excel formatado a partir de um DataFrame"""
    buffer = io.BytesIO()
    try:
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            dados_df.to_excel(
                writer,
                sheet_name=nome_aba,
                index=False,
                header=True
            )
            
            worksheet = writer.sheets[nome_aba]
            
            # Aplicar formatação completa
            num_colunas = len(dados_df.columns)
            tem_protocolo = 'Nº DE PROCESSO' in dados_df.columns
            tem_data_venc = 'DATA DE VENCIMENTO' in dados_df.columns
            tem_cpf = 'CNPJ' in dados_df.columns
            tem_modal = 'MODAL' in dados_df.columns
            
            # Ajustar larguras das colunas
            col_idx = 0
            for col in dados_df.columns:
                col_letter = chr(65 + col_idx)  # A, B, C, etc.
                if col == 'IDENTIFICADOR DE DÉBITO':
                    worksheet.column_dimensions[col_letter].width = 25
                elif col == 'Nº DE PROCESSO':
                    worksheet.column_dimensions[col_letter].width = 20
                elif col == 'DATA DE VENCIMENTO':
                    worksheet.column_dimensions[col_letter].width = 18
                elif col == 'CNPJ':
                    worksheet.column_dimensions[col_letter].width = 18
                elif col == 'MODAL':
                    worksheet.column_dimensions[col_letter].width = 18
                else:
                    worksheet.column_dimensions[col_letter].width = 15
                col_idx += 1
            
            # Formatação do cabeçalho
            header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Formatar colunas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Calcular índices das colunas
            idx_auto = None
            idx_protocolo = None
            idx_data_venc = None
            idx_cpf = None
            idx_modal = None
            
            col_idx = 1
            for col in dados_df.columns:
                if col == 'IDENTIFICADOR DE DÉBITO':
                    idx_auto = col_idx
                elif col == 'Nº DE PROCESSO':
                    idx_protocolo = col_idx
                elif col == 'DATA DE VENCIMENTO':
                    idx_data_venc = col_idx
                elif col == 'CNPJ':
                    idx_cpf = col_idx
                elif col == 'MODAL':
                    idx_modal = col_idx
                col_idx += 1
            
            # Aplicar formatação nas células
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=num_colunas):
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:  # Não formatar cabeçalho
                        if idx_cpf and cell.column == idx_cpf:  # CNPJ
                            cell.number_format = '@'
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif idx_auto and cell.column == idx_auto:  # Identificador de Débito
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        elif idx_protocolo and cell.column == idx_protocolo:  # Nº de Processo
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        elif idx_data_venc and cell.column == idx_data_venc:  # Data de Vencimento
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.number_format = '@'  # Formato texto para manter formato DD/MM/YYYY
                        elif idx_modal and cell.column == idx_modal:  # Modal
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            cell.number_format = '@'  # Formato texto
            
            worksheet.freeze_panes = 'A2'
        
        buffer.seek(0)
        excel_data = buffer.getvalue()
        return excel_data
    except Exception as e:
        buffer.close()
        raise e

# Main: carregar as duas bases
df_antiga = None
df_nova = None
if arquivo_antiga:
    with st.spinner("Carregando base antiga..."):
        df_antiga = carregar_dados(arquivo_antiga)
if arquivo_nova:
    with st.spinner("Carregando base nova..."):
        df_nova = carregar_dados(arquivo_nova)

if df_antiga is not None and df_nova is not None:
    st.markdown("---")
    st.markdown("### 📋 Preview das bases")
    col_prev1, col_prev2 = st.columns(2)
    with col_prev1:
        st.markdown("**Base antiga**")
        st.dataframe(df_antiga.head(), use_container_width=True)
        st.caption(f"Total: {len(df_antiga):,} registros")
    with col_prev2:
        st.markdown("**Base nova**")
        st.dataframe(df_nova.head(), use_container_width=True)
        st.caption(f"Total: {len(df_nova):,} registros")
    st.markdown("---")

    # Verificar colunas obrigatórias em ambas
    ok_antiga = coluna_auto in df_antiga.columns and coluna_vencimento in df_antiga.columns
    ok_nova = coluna_auto in df_nova.columns and coluna_vencimento in df_nova.columns
    if not ok_antiga:
        st.error(f"⚠️ Na base antiga: verifique as colunas '{coluna_auto}' e '{coluna_vencimento}'.")
    if not ok_nova:
        st.error(f"⚠️ Na base nova: verifique as colunas '{coluna_auto}' e '{coluna_vencimento}'.")
    if ok_antiga and ok_nova:
        if st.button("🚀 Comparar bases (antiga x nova)", type="primary", use_container_width=True):
            with st.spinner("Comparando bases por ano de vencimento..."):
                try:
                    resultado = comparar_bases(df_antiga, df_nova, coluna_auto, coluna_vencimento)
                    st.session_state['comparacao_resultado'] = resultado
                    st.session_state['coluna_auto'] = coluna_auto
                    st.session_state['coluna_vencimento'] = coluna_vencimento
                    st.session_state['coluna_cpf_cnpj'] = coluna_cpf_cnpj
                    st.session_state['coluna_modal'] = coluna_modal
                    st.session_state['coluna_protocolo'] = coluna_protocolo
                    st.session_state['nome_base_antiga'] = arquivo_antiga.name
                    st.session_state['nome_base_nova'] = arquivo_nova.name
                    st.session_state['historico_run_id'] = None  # será preenchido ao exibir resultados
                    st.success("✅ Comparação concluída!")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao comparar: {str(e)}")

# Exibir resultados da comparação
if 'comparacao_resultado' in st.session_state:
    res = st.session_state['comparacao_resultado']
    comparacao_df = res['comparacao_df']
    autos_sairam_por_ano = res['autos_sairam_por_ano']
    autos_entraram_por_ano = res['autos_entraram_por_ano']
    df_antiga_com_ano = res['df_antiga_com_ano']
    df_nova_com_ano = res['df_nova_com_ano']
    stats_antiga_por_ano = res['stats_antiga_por_ano']
    stats_nova_por_ano = res['stats_nova_por_ano']
    anos_todos = res['anos_todos']
    coluna_auto = st.session_state['coluna_auto']
    coluna_vencimento = st.session_state['coluna_vencimento']
    coluna_cpf_cnpj = st.session_state.get('coluna_cpf_cnpj', '')
    coluna_modal = st.session_state.get('coluna_modal', '')
    coluna_protocolo = st.session_state.get('coluna_protocolo', '')

    st.markdown("---")
    st.markdown("## 📊 Acompanhamento: Comparação Base Antiga x Base Nova")

    if res.get('duplicados_antiga', 0) > 0 or res.get('duplicados_nova', 0) > 0:
        st.info(
            f"ℹ️ Base antiga: {res.get('duplicados_antiga', 0):,} duplicados removidos. "
            f"Base nova: {res.get('duplicados_nova', 0):,} duplicados removidos (mantida a data de vencimento mais antiga por auto)."
        )

    # Métricas resumidas
    total_sairam = comparacao_df['Sairam'].sum()
    total_entraram = comparacao_df['Entraram'].sum()
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Anos comparados", len(anos_todos), "")
    with col2:
        st.metric("Total base antiga", f"{len(df_antiga_com_ano):,}", "")
    with col3:
        st.metric("Total base nova", f"{len(df_nova_com_ano):,}", "")
    with col4:
        st.metric("Sairam (total)", f"{total_sairam:,}", f"Entraram: {total_entraram:,}")

    # Tabela de comparação por ano
    st.markdown("### 📋 Comparação por ano de vencimento")
    st.dataframe(comparacao_df, use_container_width=True, hide_index=True)

    # Gráficos: barras antiga vs nova; barras saíram x entraram
    st.markdown("### 📈 Gráficos")
    col_a, col_b = st.columns(2)
    with col_a:
        fig_comp = go.Figure()
        fig_comp.add_trace(go.Bar(name='Base antiga', x=comparacao_df['Ano'], y=comparacao_df['Qtd base antiga'], marker_color='#1f4e79'))
        fig_comp.add_trace(go.Bar(name='Base nova', x=comparacao_df['Ano'], y=comparacao_df['Qtd base nova'], marker_color='#2e7d32'))
        fig_comp.update_layout(barmode='group', title='Quantidade: Base antiga x Base nova por ano', height=400)
        st.plotly_chart(fig_comp, use_container_width=True)
    with col_b:
        fig_mov = go.Figure()
        fig_mov.add_trace(go.Bar(name='Sairam', x=comparacao_df['Ano'], y=comparacao_df['Sairam'], marker_color='#c62828'))
        fig_mov.add_trace(go.Bar(name='Entraram', x=comparacao_df['Ano'], y=comparacao_df['Entraram'], marker_color='#1565c0'))
        fig_mov.update_layout(barmode='group', title='Sairam x Entraram por ano', height=400)
        st.plotly_chart(fig_mov, use_container_width=True)

    # Exportação: planilha de comparação + listas por ano (autos que saíram / que entraram)
    st.markdown("---")
    st.markdown("### 📥 Exportar")
    data_arquivo = datetime.now().strftime('%d %m %Y %H%M')

    # Bytes da tabela de comparação
    buf_comp = io.BytesIO()
    comparacao_df.to_excel(buf_comp, index=False, sheet_name='Comparacao')
    buf_comp.seek(0)
    excel_comparacao_bytes = buf_comp.getvalue()

    # Montar dict de Excel por ano (para download e para salvar no histórico)
    excel_por_ano_dict = {}
    for ano in sorted(anos_todos, reverse=False):
        df_ano = stats_nova_por_ano[ano]['dataframe'].copy()
        colunas_export = {}
        colunas_export['IDENTIFICADOR DE DÉBITO'] = df_ano[coluna_auto].fillna('').astype(str).str.strip()
        if coluna_protocolo and coluna_protocolo in df_ano.columns:
            colunas_export['Nº DE PROCESSO'] = df_ano[coluna_protocolo].fillna('').astype(str).str.strip()
        if coluna_modal and coluna_modal in df_ano.columns:
            colunas_export['MODAL'] = df_ano[coluna_modal].fillna('').astype(str).str.strip()
        if coluna_cpf_cnpj and coluna_cpf_cnpj in df_ano.columns:
            colunas_export['CNPJ'] = df_ano[coluna_cpf_cnpj].fillna('').astype(str).str.strip().apply(formatar_cpf_cnpj_brasileiro)
        try:
            vencimento_dt = pd.to_datetime(df_ano[coluna_vencimento], errors='coerce', dayfirst=True)
            colunas_export['DATA DE VENCIMENTO'] = vencimento_dt.dt.strftime('%d/%m/%Y').fillna('')
        except Exception:
            colunas_export['DATA DE VENCIMENTO'] = df_ano[coluna_vencimento].fillna('').astype(str).str.strip()
        ordem_colunas = ['IDENTIFICADOR DE DÉBITO']
        if 'Nº DE PROCESSO' in colunas_export:
            ordem_colunas.append('Nº DE PROCESSO')
        if 'MODAL' in colunas_export:
            ordem_colunas.append('MODAL')
        if 'CNPJ' in colunas_export:
            ordem_colunas.append('CNPJ')
        ordem_colunas.append('DATA DE VENCIMENTO')
        dados_exportacao = pd.DataFrame({col: colunas_export[col] for col in ordem_colunas}, index=df_ano.index)
        dados_exportacao = dados_exportacao.dropna(how='all')
        if not dados_exportacao.empty:
            try:
                excel_por_ano_dict[ano] = gerar_excel_formatado(
                    dados_exportacao,
                    f'Autos_{ano}',
                    f'Autos Vencimento {ano}.xlsx'
                )
            except Exception:
                pass

    # Salvar no histórico (SQLite + pasta) na primeira exibição desta comparação
    if st.session_state.get('historico_run_id') is None:
        try:
            config = {
                'coluna_auto': coluna_auto,
                'coluna_vencimento': coluna_vencimento,
                'coluna_protocolo': coluna_protocolo,
                'coluna_modal': coluna_modal,
                'coluna_cpf_cnpj': coluna_cpf_cnpj,
            }
            nome_antiga = st.session_state.get('nome_base_antiga') or 'base_antiga'
            nome_nova = st.session_state.get('nome_base_nova') or 'base_nova'
            run_id = save_run(
                res,
                nome_antiga,
                nome_nova,
                config,
                excel_comparacao_bytes,
                excel_por_ano_dict,
            )
            st.session_state['historico_run_id'] = run_id
            st.success("✅ Comparação salva no histórico. Você pode acessá-la na seção **Histórico de comparações** abaixo.")
        except Exception as e:
            st.warning(f"Não foi possível salvar no histórico: {e}")

    # Download da tabela de comparação
    st.download_button(
        label="📥 Download tabela de comparação (Excel)",
        data=excel_comparacao_bytes,
        file_name=f"Comparacao_bases_{data_arquivo}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="download_comparacao"
    )

    # Exportação por ano de vencimento (base nova)
    st.markdown("#### 📥 Exportar por ano de vencimento (base nova)")
    st.info("💡 Baixe planilhas separadas por ano de vencimento. Cada arquivo contém todos os autos da base nova com vencimento naquele ano (do ano mais antigo ao mais novo).")
    num_colunas_export = min(3, len(anos_todos))
    cols_export = st.columns(num_colunas_export)
    for idx, ano in enumerate(sorted(anos_todos, reverse=False)):
        col_idx = idx % num_colunas_export
        with cols_export[col_idx]:
            st.markdown(f"##### 📅 Ano {ano}")
            qtd_autos = stats_nova_por_ano[ano]['quantidade']
            if ano in excel_por_ano_dict:
                st.download_button(
                    label=f"📥 Download Autos {ano}",
                    data=excel_por_ano_dict[ano],
                    file_name=f"Autos Vencimento {ano} {data_arquivo}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=f"download_ano_{ano}",
                    help=f"Arquivo Excel com todos os autos de vencimento {ano} (base nova)"
                )
                st.caption(f"✅ {qtd_autos:,} autos")
            else:
                st.caption(f"📊 Ano {ano}: 0 autos")

    st.markdown("---")
    st.caption(f"📅 Data da comparação: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    st.caption(f"📊 Anos: {', '.join([str(a) for a in sorted(anos_todos)])}")

else:
    st.info("👆 Faça o upload da **base antiga** e da **base nova** na barra lateral e clique em **Comparar bases**.")
    
    with st.expander("ℹ️ Como usar o sistema"):
        st.markdown("""
        ### Instruções de Uso:
        
        1. **Upload das bases**: Na barra lateral, envie duas planilhas:
           - **Base antiga** (período anterior)
           - **Base nova** (período atual)
        2. **Configuração de colunas**: Informe os nomes exatos das colunas (valem para as duas bases):
           - ⚠️ **Coluna Auto de Infração** (OBRIGATÓRIA)
           - ⚠️ **Coluna Vencimento** (OBRIGATÓRIA)
           - Coluna Nº do Processo, Modal, CPF/CNPJ (Opcionais)
        3. **Comparar**: Clique em **Comparar bases (antiga x nova)**.
        4. **Resultados**: Veja por ano de vencimento:
           - **Qtd base antiga** e **Qtd base nova**
           - **Sairam**: autos que estavam na base antiga naquele ano e não estão na base nova
           - **Entraram**: autos que estão na base nova naquele ano e não estavam na base antiga
        5. **Exportar**: Baixe a tabela de comparação e, por ano de vencimento, um Excel com todos os autos da base nova (do ano mais antigo ao mais novo).
        
        ### Funcionalidades:
        - ✅ Comparação de duas bases por ano de vencimento
        - ✅ Tabela e gráficos: antiga x nova e saíram x entraram
        - ✅ Exportação por ano de vencimento (um arquivo por ano, base nova)
        - ✅ Formatação profissional das planilhas Excel
        - ✅ Histórico de comparações salvo em SQLite + pasta (veja abaixo)
        """)

# --- Histórico de comparações (sempre visível) ---
st.markdown("---")
st.markdown("## 📚 Histórico de comparações")
init_db()
runs = list_runs()
if not runs:
    st.info("Nenhuma comparação salva ainda. Após comparar duas bases, o resultado será salvo aqui automaticamente.")
else:
    for r in runs:
        with st.expander(
            f"🕐 {r['data_hora']} — **{r['nome_base_antiga']}** × **{r['nome_base_nova']}** "
            f"(antiga: {r['total_antiga']:,} | nova: {r['total_nova']:,} | saíram: {r['total_sairam']:,} | entraram: {r['total_entraram']:,})"
        ):
            run_full = get_run(r["id"])
            if not run_full:
                st.caption("Registro não encontrado.")
                continue
            st.markdown(f"**Base antiga:** {run_full['nome_base_antiga']} | **Base nova:** {run_full['nome_base_nova']}")
            st.dataframe(run_full["comparacao_df"], use_container_width=True, hide_index=True)
            col_a, col_b = st.columns(2)
            with col_a:
                # Download ZIP com todos os arquivos da pasta
                if run_full.get("arquivos"):
                    buf_zip = io.BytesIO()
                    with zipfile.ZipFile(buf_zip, "w", zipfile.ZIP_DEFLATED) as zf:
                        for f in run_full["arquivos"]:
                            if f.is_file():
                                zf.write(str(f), f.name)
                    buf_zip.seek(0)
                    # Nome do ZIP = nomes das bases (sem extensão, caracteres seguros)
                    def nome_seguro(s):
                        if not s:
                            return "base"
                        n = Path(s).stem
                        n = re.sub(r'[<>:"/\\|?*]', "_", n).strip() or "base"
                        return n[:80]
                    nome_zip = f"{nome_seguro(run_full['nome_base_antiga'])} x {nome_seguro(run_full['nome_base_nova'])}.zip"
                    st.download_button(
                        label="📥 Baixar todos os arquivos (ZIP)",
                        data=buf_zip,
                        file_name=nome_zip,
                        mime="application/zip",
                        key=f"zip_{r['id']}",
                    )
            with col_b:
                if st.button("🗑️ Excluir deste histórico", key=f"del_{r['id']}"):
                    excluir_run(r["id"])
                    st.rerun()

