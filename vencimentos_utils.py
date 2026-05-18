"""Funções de apoio do app de comparação por ano."""

import io

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


@st.cache_data
def carregar_dados_vencimentos(arquivo):
    """Carrega a base e já tira as linhas em branco que só atrapalham a leitura."""
    try:
        if arquivo.name.endswith(".csv"):
            df = pd.read_csv(arquivo, encoding="utf-8", sep=";", decimal=",", header=0)
        else:
            df = pd.read_excel(arquivo, header=0)
        return df.dropna(how="all")
    except Exception as exc:
        st.error(f"Erro ao carregar arquivo: {exc}")
        return None


def _serie_vencimento_datetime(df, coluna_vencimento):
    """Converte a coluna de vencimento para data de trabalho do pandas."""
    if df[coluna_vencimento].dtype == "datetime64[ns]":
        return df[coluna_vencimento]
    return pd.to_datetime(
        df[coluna_vencimento],
        errors="coerce",
        dayfirst=True,
    )


def remover_duplicados_manter_mais_antiga(df, coluna_auto, coluna_vencimento):
    """Se o mesmo auto aparecer mais de uma vez, fica o vencimento mais antigo."""
    df_resultado = df.copy()
    try:
        df_resultado["_VENCIMENTO_DT"] = _serie_vencimento_datetime(df_resultado, coluna_vencimento)
        df_resultado = df_resultado.sort_values(
            by=[coluna_auto, "_VENCIMENTO_DT"],
            ascending=[True, True],
            na_position="last",
        )
        return df_resultado.drop_duplicates(subset=[coluna_auto], keep="first").drop(columns=["_VENCIMENTO_DT"])
    except Exception as exc:
        st.error(f"Erro ao remover duplicados: {exc}")
        return df


def extrair_ano_vencimento(df, coluna_vencimento):
    """Cria a coluna de ano para o restante das comparações."""
    df_resultado = df.copy()
    try:
        df_resultado["_VENCIMENTO_DT"] = _serie_vencimento_datetime(df_resultado, coluna_vencimento)
        df_resultado["ANO_VENCIMENTO"] = df_resultado["_VENCIMENTO_DT"].dt.year
        return df_resultado.drop(columns=["_VENCIMENTO_DT"])
    except Exception as exc:
        st.error(f"Erro ao extrair ano: {exc}")
        return df


def comparar_bases(df_antiga, df_nova, coluna_auto, coluna_vencimento):
    """Compara base antiga e base nova olhando o ano de vencimento."""
    df_antiga_limpa = remover_duplicados_manter_mais_antiga(df_antiga, coluna_auto, coluna_vencimento)
    df_antiga_com_ano = extrair_ano_vencimento(df_antiga_limpa, coluna_vencimento)
    df_antiga_com_ano = df_antiga_com_ano[df_antiga_com_ano["ANO_VENCIMENTO"].notna()].copy()
    df_antiga_com_ano["ANO_VENCIMENTO"] = df_antiga_com_ano["ANO_VENCIMENTO"].astype(int)

    df_nova_limpa = remover_duplicados_manter_mais_antiga(df_nova, coluna_auto, coluna_vencimento)
    df_nova_com_ano = extrair_ano_vencimento(df_nova_limpa, coluna_vencimento)
    df_nova_com_ano = df_nova_com_ano[df_nova_com_ano["ANO_VENCIMENTO"].notna()].copy()
    df_nova_com_ano["ANO_VENCIMENTO"] = df_nova_com_ano["ANO_VENCIMENTO"].astype(int)

    def autos_por_ano(df_com_ano):
        # Aqui vira um mapa simples de ano -> autos, que deixa a comparação mais direta.
        resultado = {}
        for ano, grupo in df_com_ano.groupby("ANO_VENCIMENTO"):
            resultado[int(ano)] = set(grupo[coluna_auto].astype(str).str.strip().values)
        return resultado

    antiga_por_ano = autos_por_ano(df_antiga_com_ano)
    nova_por_ano = autos_por_ano(df_nova_com_ano)
    anos_todos = sorted(set(antiga_por_ano.keys()) | set(nova_por_ano.keys()))

    stats_antiga_por_ano = {}
    stats_nova_por_ano = {}
    for ano in anos_todos:
        df_ano_antiga = df_antiga_com_ano[df_antiga_com_ano["ANO_VENCIMENTO"] == ano]
        df_ano_nova = df_nova_com_ano[df_nova_com_ano["ANO_VENCIMENTO"] == ano]
        stats_antiga_por_ano[ano] = {"quantidade": len(df_ano_antiga), "dataframe": df_ano_antiga}
        stats_nova_por_ano[ano] = {"quantidade": len(df_ano_nova), "dataframe": df_ano_nova}

    linhas = []
    autos_sairam_por_ano = {}
    autos_entraram_por_ano = {}
    for ano in anos_todos:
        set_antiga = antiga_por_ano.get(ano, set())
        set_nova = nova_por_ano.get(ano, set())
        sairam = set_antiga - set_nova
        entraram = set_nova - set_antiga
        autos_sairam_por_ano[ano] = sairam
        autos_entraram_por_ano[ano] = entraram
        linhas.append(
            {
                "Ano": ano,
                "Qtd base antiga": len(set_antiga),
                "Qtd base nova": len(set_nova),
                "Sairam": len(sairam),
                "Entraram": len(entraram),
            }
        )

    return {
        "comparacao_df": pd.DataFrame(linhas),
        "autos_sairam_por_ano": autos_sairam_por_ano,
        "autos_entraram_por_ano": autos_entraram_por_ano,
        "df_antiga_com_ano": df_antiga_com_ano,
        "df_nova_com_ano": df_nova_com_ano,
        "stats_antiga_por_ano": stats_antiga_por_ano,
        "stats_nova_por_ano": stats_nova_por_ano,
        "anos_todos": anos_todos,
        "duplicados_antiga": len(df_antiga) - len(df_antiga_limpa),
        "duplicados_nova": len(df_nova) - len(df_nova_limpa),
    }


def formatar_cpf_cnpj_brasileiro(valor):
    """Formata CPF ou CNPJ para a exportação ficar mais legível."""
    if pd.isna(valor) or valor == "" or valor is None:
        return ""
    valor_str = str(valor).replace(".", "").replace("-", "").replace("/", "").strip()
    if not valor_str or not valor_str.isdigit():
        return str(valor)
    if len(valor_str) == 11:
        return f"{valor_str[0:3]}.{valor_str[3:6]}.{valor_str[6:9]}-{valor_str[9:11]}"
    if len(valor_str) == 14:
        return f"{valor_str[0:2]}.{valor_str[2:5]}.{valor_str[5:8]}/{valor_str[8:12]}-{valor_str[12:14]}"
    return str(valor)


def gerar_excel_vencimentos_formatado(dados_df, nome_aba, nome_arquivo):
    """Monta o Excel no padrão visual que o time já usa nesse relatório."""
    del nome_arquivo
    buffer = io.BytesIO()
    try:
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            dados_df.to_excel(writer, sheet_name=nome_aba, index=False, header=True)
            worksheet = writer.sheets[nome_aba]
            num_colunas = len(dados_df.columns)

            for col_idx, col in enumerate(dados_df.columns):
                col_letter = chr(65 + col_idx)
                if col == "IDENTIFICADOR DE DÉBITO":
                    worksheet.column_dimensions[col_letter].width = 25
                elif col == "Nº DE PROCESSO":
                    worksheet.column_dimensions[col_letter].width = 20
                elif col in {"DATA DE VENCIMENTO", "CNPJ", "MODAL"}:
                    worksheet.column_dimensions[col_letter].width = 18
                else:
                    worksheet.column_dimensions[col_letter].width = 15

            header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            idx_por_coluna = {nome: idx + 1 for idx, nome in enumerate(dados_df.columns)}
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=num_colunas):
                for cell in row:
                    cell.border = thin_border
                    if cell.row <= 1:
                        continue
                    if cell.column == idx_por_coluna.get("CNPJ"):
                        cell.number_format = "@"
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif cell.column == idx_por_coluna.get("IDENTIFICADOR DE DÉBITO"):
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif cell.column == idx_por_coluna.get("Nº DE PROCESSO"):
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    elif cell.column == idx_por_coluna.get("DATA DE VENCIMENTO"):
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.number_format = "@"
                    elif cell.column == idx_por_coluna.get("MODAL"):
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        cell.number_format = "@"

            worksheet.freeze_panes = "A2"

        buffer.seek(0)
        return buffer.getvalue()
    except Exception:
        buffer.close()
        raise
